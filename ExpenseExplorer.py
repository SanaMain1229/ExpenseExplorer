import sys

import pandas as pd
import openpyxl
import os

from openpyxl import Workbook

from PyQt5.QtCore import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.uic import *
from PyQt5.QtWidgets import QApplication


from Home_ui import Ui_Home
from AddExpense_ui import Ui_AddData

class GetCalendar(QWidget):
    def __init__(self, getLE, getMain):
        super().__init__()
        self.currLE = getLE
        self.currMain = getMain
        self.calendar = QCalendarWidget(self)
        self.calendar.clicked.connect(self.selectDate)

        layout = QVBoxLayout()
        layout.addWidget(self.calendar)
        self.setLayout(layout)

        self.setWindowFlags(Qt.Popup)

    def selectDate(self, date):
        self.currLE.setText(date.toString("yyyy-MM-dd"))
        #self.currMain.transactData()
        self.close()
    
        

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.Home = Ui_Home()
        self.Home.setupUi(self)
        self.setFixedSize(self.size())

        self.Home.leSDay.setText(QDate.currentDate().toString('yyyy-MM-dd'))
        self.Home.leEDay.setText(QDate.currentDate().toString('yyyy-MM-dd'))

        self.SDay = self.Home.leSDay.text()
        self.EDay = self.Home.leEDay.text()

        curDir = os.path.dirname(os.path.abspath(__file__))
        FileName = "ExpenseWorkbook.xlsx"
        self.FilePath = os.path.join(curDir, FileName)

        self.wb = openpyxl.load_workbook(self.FilePath)
        self.ws = self.wb["Expense Tracker"]

        self.Home.leSDay.mousePressEvent = lambda event, le = self.Home.leSDay: self.addDate(event, le)
        self.Home.leEDay.mousePressEvent = lambda event, le = self.Home.leEDay: self.addDate(event, le)
        self.Home.leSDay.textChanged.connect(self.transactData)
        self.Home.leEDay.textChanged.connect(self.transactData)
        self.transactData()
        self.Home.actionHAdd.triggered.connect(self.addAction)

        self.Home.checkC.stateChanged.connect(self.transactData)
        self.Home.checkF.stateChanged.connect(self.transactData)
        self.Home.checkG.stateChanged.connect(self.transactData)
        self.Home.checkO.stateChanged.connect(self.transactData)
        self.Home.checkP.stateChanged.connect(self.transactData)
        self.Home.checkS.stateChanged.connect(self.transactData)
        
    def addDate(self, event, le):
        self.calendar_popup = GetCalendar(le, self)
        rect = le.geometry()
        global_position = self.mapToGlobal(rect.bottomLeft())
        self.calendar_popup.setGeometry(global_position.x(), global_position.y(), 300, 200)  # Adjust position and size
        self.calendar_popup.show()
        
    def transactData(self):
        PClothes = 0
        PFood = 0
        PBills = 0
        PSports = 0
        PGifts = 0
        POthers = 0
        self.Home.twSummary.setRowCount(0)
        PCont = False
        for row in range(2, 10000):

            if self.ws[f'A{row}'].value is None:
                #₱
                self.Home.lblC.setText(f"CLOTHES\n₱ {PClothes}")
                self.Home.lblF.setText(f"FOODS\n₱ {PFood}")
                self.Home.lblG.setText(f"GIFTS\n₱ {PGifts}")
                self.Home.lblO.setText(f"OTHERS\n₱ {POthers}")
                self.Home.lblP.setText(f"PAYMENTS/BILLS\n₱ {PBills}")
                self.Home.lblS.setText(f"SPORTS\n₱ {PSports}")
                break
            SDate = QDate.fromString(self.Home.leSDay.text(), "yyyy-MM-dd")
            EDate = QDate.fromString(self.Home.leEDay.text(), "yyyy-MM-dd")
            
            CDate = QDate.fromString(self.ws[f'A{row}'].value, "yyyy-MM-dd")

            if SDate > EDate:
                TDate = SDate
                SDate = EDate
                EDate = TDate
            

            if self.ws[f'D{row}'].value == "CLOTHES" and self.Home.checkC.isChecked() and SDate <= CDate and EDate >= CDate:
                PClothes = PClothes + float(self.ws[f'C{row}'].value)
                PCont = True

            if self.ws[f'D{row}'].value == "FOODS" and self.Home.checkF.isChecked() and SDate <= CDate and EDate >= CDate:
                PFood = PFood + float(self.ws[f'C{row}'].value)
                PCont = True

            if self.ws[f'D{row}'].value == "PAYMENTS/BILLS" and self.Home.checkP.isChecked() and SDate <= CDate and EDate >= CDate:
                PBills = PBills + float(self.ws[f'C{row}'].value)
                PCont = True

            if self.ws[f'D{row}'].value == "SPORTS" and self.Home.checkS.isChecked() and SDate <= CDate and EDate >= CDate:
                PSports = PSports + float(self.ws[f'C{row}'].value)
                PCont = True

            if self.ws[f'D{row}'].value == "GIFTS" and self.Home.checkG.isChecked() and SDate <= CDate and EDate >= CDate:
                PGifts = PGifts + float(self.ws[f'C{row}'].value)
                PCont = True

            if self.ws[f'D{row}'].value == "OTHERS" and self.Home.checkO.isChecked() and SDate <= CDate and EDate >= CDate:
                POthers = POthers + float(self.ws[f'C{row}'].value)
                PCont = True



            if PCont:
                self.rpos = self.Home.twSummary.rowCount()
                self.Home.twSummary.insertRow(self.rpos)
                PCont = False

                for col in range(1,5):
                    item = QTableWidgetItem(str(self.ws.cell(row=row, column=col).value))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                    if col == 3:
                        item.setText(f"₱ {item.text()}")

                    self.Home.twSummary.setItem(self.rpos, col-1, item)
                
                self.Home.twSummary.resizeColumnsToContents()

            
            #for col in range (self.Home.twSummary.columnCount()):
            #    cur_wid = self.Home.twSummary.columnWidth(col)
            #    self.Home.twSummary.setColumnWidth(col, cur_wid)
                
    def addAction(self):

        self.AddFrm = QDialog(self)
        self.AddUI = Ui_AddData()
        self.AddUI.setupUi(self.AddFrm)

        self.AddFrm.show()
        self.AddUI.leSDate.mousePressEvent = lambda event, le = self.AddUI.leSDate: self.addDate(event, le)
        self.AddUI.pbAdd.clicked.connect(self.AddExpenses)

    def AddExpenses(self):
         
        if self.AddUI.leSDate.text().strip() == "":
            QMessageBox.critical(self.AddFrm, "Date is Empty", "Select a date to proceed!")
            return
        
        if self.AddUI.leISpend.text().strip() == "":
            QMessageBox.critical(self.AddFrm, "Item is Empty", "Input your spended item to proceed!")
            return
        
        if self.AddUI.leIPrice.text().strip() == "":
            QMessageBox.critical(self.AddFrm, "Price is Empty", "Input the price of the item to proceed!")
            return
        
        if self.AddUI.comboCat.currentText() == "":
            QMessageBox.critical(self.AddFrm, "Category is Empty", "Select a category of the item to proceed!")
            return
        
        try:
            self.CPrice = int(self.AddUI.leIPrice.text().strip())

        except ValueError as e:
            QMessageBox.critical(self.AddFrm, "Invalid Price", "Recheck the Item Pricing")
            return

        for row in range(2, 10000):
            
            if self.ws[f'A{row}'].value is None:

                self.ws[f'A{row}'].value = self.AddUI.leSDate.text()
                self.ws[f'B{row}'].value = self.AddUI.leISpend.text()
                self.ws[f'C{row}'].value = self.AddUI.leIPrice.text()
                self.ws[f'D{row}'].value = self.AddUI.comboCat.currentText()

                self.wb.save(self.FilePath)
                QMessageBox.information(self.AddFrm, "Expense Added", f"Item {self.AddUI.leISpend.text()} is now added")
                self.AddUI.leISpend.clear()
                self.AddUI.leIPrice.clear()
                self.AddUI.comboCat.setCurrentIndex(-1)
                self.transactData()
                break

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())